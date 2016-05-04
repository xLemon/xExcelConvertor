#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import absolute_import
from __future__ import unicode_literals

import traceback

from definitions.constant_data import xConstantData

from utilities.excel_helper import xExcelHelper
from utilities.export_helper import xExportHelper

from core.processor_manager import xProcessorManager

from openpyxl.reader.excel import load_workbook

class xWorksheetConvertor :
	@staticmethod
	def ConvertWorksheet(p_strWorkbookName, p_cWorkbook, p_mapExportConfigs, p_mapDatabaseConfigs, p_mapIndexSheetConfigs) :
		cWorkSheet = p_cWorkbook.get_sheet_by_name(p_mapIndexSheetConfigs['DATA_SHEET'])

		if cWorkSheet is None :
			print('>>>>> 错误：没有找到工作表 : {0}'.format(p_mapIndexSheetConfigs['DATA_SHEET']))
			return False

		if not xWorksheetConvertor.__IsNeedExport(cWorkSheet) :
			print('>>>>> 工作表 [{0}] 没有指定导出字段，跳过 ...'.format(p_mapIndexSheetConfigs['DATA_SHEET']))
			return True

		print('>>>>> 正在验证 工作表 [{0}] 表头配置 ...'.format(p_mapIndexSheetConfigs['DATA_SHEET']))

		bSuccess, mapDataSheetConfigs = xWorksheetConvertor.__GetDataSheetTitleDictionary(cWorkSheet)

		if not bSuccess :
			return False

		if not xWorksheetConvertor.__ValidateDataSheetTitleMap(cWorkSheet, mapDataSheetConfigs) :
			return False

		print('>>>>> 工作表 [{0}] 表头配置验证成功！'.format(p_mapIndexSheetConfigs['DATA_SHEET']))

		bSuccess = True

		mapProcessorInstances = {}

		for strExportType in p_mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES'] :
			cDataSheetDatas = xWorksheetConvertor.__GetDataSheetDataGenerator(cWorkSheet, mapDataSheetConfigs.keys()[0], mapDataSheetConfigs.keys()[-1])

			bSuccess, nCategoryLevel, mapPreloadDataMaps = xWorksheetConvertor.__GetPreloadDataMaps(cWorkSheet, p_mapExportConfigs, mapDataSheetConfigs, cDataSheetDatas, strExportType)

			if not bSuccess :
				break

			bSuccess = False

			if mapProcessorInstances.get(strExportType) is None :
				mapProcessorInstances[strExportType] = xProcessorManager.GetProcessorInstance(strExportType)
				
			if mapProcessorInstances[strExportType] is None :
				break

			if not mapProcessorInstances[strExportType].ProcessExport(p_strWorkbookName, p_cWorkbook, cWorkSheet, p_mapExportConfigs, p_mapDatabaseConfigs, p_mapIndexSheetConfigs, mapDataSheetConfigs, mapPreloadDataMaps, nCategoryLevel) :
				break

			bSuccess = True

		return bSuccess

	@staticmethod
	def __ValidateDataSheetTitleMap(p_cWorkSheet, p_mapDataSheetConfigs) :
		bSuccess = True

		for nColumnIndex in p_mapDataSheetConfigs :
			bSuccess = False

			if p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_FIELD] is None :
				print('>>>>> 错误：工作表 [{0}] [{1}{2}] 单元格字段名未配置！'.format(p_cWorkSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(nColumnIndex - 1), xConstantData.DATA_SHEET_ROW_FIELD))
				break

			if p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_TYPE] is None :
				print('>>>>> 错误：工作表 [{0}] [{1}{2}] 单元格 [{3}] 字段 数据类型未配置！'.format(p_cWorkSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(nColumnIndex - 1), xConstantData.DATA_SHEET_ROW_DATA_TYPE, p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_FIELD]))
				break

			if (p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_LENGTH] is None or p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_LENGTH] <= 0) and xConstantData.MYSQL_DATA_DEFINITIONS[p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_TYPE].upper()]['LENGTH'] > 0 :
				# print('>>>>> 提示：工作表 [{0}] [{1}{2}] 单元格 [{3}] 字段 数据类型 [{4}] 长度未配置或配置错误，使用默认长度 [%d] ...'.format(p_cWorkSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(nColumnIndex - 1), xConstantData.DATA_SHEET_ROW_DATA_TYPE, p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_FIELD], p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_TYPE].upper(), xConstantData.MYSQL_DATA_DEFINITIONS[p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_TYPE].upper()]['LENGTH']))
				p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_LENGTH] = xConstantData.MYSQL_DATA_DEFINITIONS[p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DATA_TYPE].upper()]['LENGTH']

			if p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_AUTO_INCREMENT_IDENTIFIER] is not None and p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DEFAULT_VALUE] is not None :
				print('>>>>> 错误：工作表 [{0}] 自增字段 [{1}] 不允许配置默认值！' % (p_cWorkSheet.title, p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_FIELD]))
				break

			if p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_UNSIGNED_IDENTIFIER] is not None and p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DEFAULT_VALUE] is not None and p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_DEFAULT_VALUE] < 0 :
				print('>>>>> 错误：工作表 [{0}] [{1}{2}] 单元格 [{3}] 字段 为无符号数据，其默认值不允许配置负数！'.format(p_cWorkSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(nColumnIndex - 1), xConstantData.DATA_SHEET_ROW_DATA_TYPE, p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_FIELD]))
				break

			bSuccess = True

		return bSuccess

	@staticmethod
	def __IsNeedExport(p_cWorkSheet) :
		nMaxColumns = p_cWorkSheet.max_column

		if nMaxColumns <= 0 :
			return False

		nNonExportCount = 0

		cExportIdentifiers = p_cWorkSheet.get_squared_range(2, xConstantData.DATA_SHEET_ROW_EXPORT_IDENTIFIER, nMaxColumns, xConstantData.DATA_SHEET_ROW_EXPORT_IDENTIFIER)
		
		for cExportIdentifierCells in cExportIdentifiers :
			for cExportIdentifierCell in cExportIdentifierCells :
				if cExportIdentifierCell.value is not None and 0 == cmp(cExportIdentifierCell.value, xConstantData.EXPORT_IDENTIFIER_NONE) :
					nNonExportCount += 1

		if nNonExportCount >= (nMaxColumns - 1) :
			return False

		return True

	@staticmethod
	def __GetDataSheetTitleDictionary(p_cWorkSheet) :
		mapDataSheetTitles = {}

		nHighestColumnIndex = p_cWorkSheet.max_column
		
		if nHighestColumnIndex <= 1 :
			return mapDataSheetTitles

		# 获取所有列头数据
		
		for nColumnIndex in range(2, nHighestColumnIndex + 1) :
			mapDataSheetTitles[nColumnIndex] = {}

			cDataSheetTitles = p_cWorkSheet.get_squared_range(nColumnIndex, xConstantData.DATA_SHEET_ROW_MIN, nColumnIndex, xConstantData.DATA_SHEET_ROW_MAX)

			nRowIndex = xConstantData.DATA_SHEET_ROW_MIN

			for cTitleCells in cDataSheetTitles :
				mapDataSheetTitles[nColumnIndex][nRowIndex] = cTitleCells[0].value
				nRowIndex += 1

		# 删除无效列头数据

		bSuccess = True

		nLastEffectiveColumnIndex = -1

		for nColumnIndex in range(2, nHighestColumnIndex + 1)[::-1] :
			lstInvalidRowIndexs = []

			for nRowIndex in range(xConstantData.DATA_SHEET_ROW_MIN, xConstantData.DATA_SHEET_ROW_MAX + 1) :
				if mapDataSheetTitles[nColumnIndex][nRowIndex] is None :
					lstInvalidRowIndexs.append(nRowIndex)

			if len(lstInvalidRowIndexs) == (xConstantData.DATA_SHEET_ROW_MAX - xConstantData.DATA_SHEET_ROW_MIN + 1) :
				if nLastEffectiveColumnIndex < 0 :
					del(mapDataSheetTitles[nColumnIndex])
				else :
					print('>>>>> 错误：工作表 [{0}] [{1}] 列 表头未配置！'.format(p_cWorkSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(nColumnIndex - 1)))
					bSuccess = False
					break
			else :
				nLastEffectiveColumnIndex = nColumnIndex

		if not bSuccess :
			mapDataSheetTitles.clear()

		return bSuccess, mapDataSheetTitles

	@staticmethod
	def __GetDataSheetDataGenerator(p_cWorkSheet, p_nMinColumn, p_nMaxColumn) :
		lstEffectiveColumnIndexs = xWorksheetConvertor.__GetEffectiveColumnIndexs(p_cWorkSheet)

		nEffectiveColumns = len(lstEffectiveColumnIndexs)

		if nEffectiveColumns <= 0 :
			return None

		cDataSheetDatas = p_cWorkSheet.get_squared_range(p_nMinColumn, xConstantData.DATA_SHEET_ROW_MAX + 1, p_nMaxColumn, p_cWorkSheet.max_row)

		return cDataSheetDatas

	@staticmethod
	def __GetEffectiveColumnIndexs(p_cWorkSheet) :
		lstEffectiveColumnIndexs = []

		nColumnIndex = 1

		cKeyCellRows = p_cWorkSheet.get_squared_range(2, xConstantData.DATA_SHEET_ROW_FIELD, p_cWorkSheet.max_column, xConstantData.DATA_SHEET_ROW_FIELD)
		
		for cKeyCells in cKeyCellRows :
			for cKeyCell in cKeyCells :
				if cKeyCell.value is not None :
					lstEffectiveColumnIndexs.append(nColumnIndex + 1)
				nColumnIndex += 1

		return lstEffectiveColumnIndexs

	@staticmethod
	def __GetCategoryLevel(p_cWorkSheet) :
		if p_cWorkSheet.max_column <= 0 :
			return 0

		nCategoryLevel = 0

		cDataSheetTireCellRows = p_cWorkSheet.get_squared_range(xConstantData.DATA_SHEET_NAVIGATION_COLUMN_CATEGORY_LEVEL, xConstantData.DATA_SHEET_ROW_NAVIGATION, xConstantData.DATA_SHEET_NAVIGATION_COLUMN_CATEGORY_LEVEL, xConstantData.DATA_SHEET_ROW_NAVIGATION)

		for cDataSheetTireCells in cDataSheetTireCellRows :
			if cDataSheetTireCells[0].value is None :
				nCategoryLevel = 0
			else :
				nCategoryLevel = cDataSheetTireCells[0].value

		return nCategoryLevel

	@staticmethod
	def __GetAvaiableDataSheetConfigCount(p_strExportType, p_mapExportConfigs, p_mapDataSheetConfigs) :
		nAvaiableDataSheetConfigCount = 0

		for nColumnIndex in p_mapDataSheetConfigs :
			if not xExportHelper.IsDataSheetColumnLanguageAvailable(p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_LANGUAGE_CODE], p_strExportType, p_mapExportConfigs) :
				continue

			if not xExportHelper.IsDataSheetColumnExportTypeAvailable(p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_EXPORT_IDENTIFIER], p_strExportType, p_mapExportConfigs) :
				continue

			nAvaiableDataSheetConfigCount += 1

		return nAvaiableDataSheetConfigCount

	@staticmethod
	def __GetCategoryLevelDataContainer(p_mixContainer, p_lstCategoryLevelValues = [], p_nCategoryLevelIndex = 0, p_bFirst = True) :
		if p_nCategoryLevelIndex < 0 :
			p_nCategoryLevelIndex = 0

		nMaxCategoryLevels = len(p_lstCategoryLevelValues)

		if 0 == p_nCategoryLevelIndex and p_bFirst :
			if type(p_mixContainer) == dict and not p_mixContainer.has_key('datas') :
				if 0 == nMaxCategoryLevels :
					p_mixContainer['datas'] = []
				else :
					p_mixContainer['datas'] = {}
			return xWorksheetConvertor.__GetCategoryLevelDataContainer(p_mixContainer['datas'], p_lstCategoryLevelValues, p_nCategoryLevelIndex, False)
		
		if nMaxCategoryLevels > 0 and p_nCategoryLevelIndex < nMaxCategoryLevels and not p_mixContainer.has_key(p_lstCategoryLevelValues[p_nCategoryLevelIndex]) :
			if p_nCategoryLevelIndex == nMaxCategoryLevels - 1 :
				p_mixContainer[p_lstCategoryLevelValues[p_nCategoryLevelIndex]] = []
			else :
				p_mixContainer[p_lstCategoryLevelValues[p_nCategoryLevelIndex]] = {}
		
		if p_nCategoryLevelIndex >= nMaxCategoryLevels :
			return p_mixContainer

		return xWorksheetConvertor.__GetCategoryLevelDataContainer(p_mixContainer[p_lstCategoryLevelValues[p_nCategoryLevelIndex]], p_lstCategoryLevelValues, p_nCategoryLevelIndex + 1, False)

	@staticmethod
	def __GetCategoryLevelCellIndexList(p_strExportType, p_nCategoryLevel, p_mapExportConfigs, p_mapDataSheetConfigs) :
		if p_nCategoryLevel <= 0 :
			return []

		lstCategoryLevelCellIndexs = []

		nCategoryLevelIndex = 0

		nCellIndex = 0

		for nColumnIndex in p_mapDataSheetConfigs :
			if nCategoryLevelIndex >= p_nCategoryLevel :
				break

			if not xExportHelper.IsDataSheetColumnLanguageAvailable(p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_LANGUAGE_CODE], p_strExportType, p_mapExportConfigs) :
				nCellIndex += 1
				continue

			if not xExportHelper.IsDataSheetColumnExportTypeAvailable(p_mapDataSheetConfigs[nColumnIndex][xConstantData.DATA_SHEET_ROW_EXPORT_IDENTIFIER], p_strExportType, p_mapExportConfigs) :
				nCellIndex += 1
				continue

			lstCategoryLevelCellIndexs.append(nCellIndex)

			nCellIndex += 1
			
			nCategoryLevelIndex += 1

		return lstCategoryLevelCellIndexs

	@staticmethod
	def __GetPreloadDataMaps(p_cWorkSheet, p_mapExportConfigs, p_mapDataSheetConfigs, p_cDataSheetDatas, p_strExportType) :
		nCategoryLevel = xWorksheetConvertor.__GetCategoryLevel(p_cWorkSheet)

		nAvaiableDataSheetConfigCount = xWorksheetConvertor.__GetAvaiableDataSheetConfigCount(p_strExportType, p_mapExportConfigs, p_mapDataSheetConfigs)
		
		if nCategoryLevel > nAvaiableDataSheetConfigCount :
			print('>>>>> 错误：工作表 [{0}] 导出类型 [{1}] 分类层级配置错误，不能超过有效表头配置列数！'.format(p_cWorkSheet.title, p_strExportType))
			return False, nCategoryLevel, {}

		print('>>>>> 工作表 [{0}] 导出类型 [{1}] 分类层级 : {2}'.format(p_cWorkSheet.title, p_strExportType, nCategoryLevel))

		lstCategoryLevelIndexs = xWorksheetConvertor.__GetCategoryLevelCellIndexList(p_strExportType, nCategoryLevel, p_mapDataSheetConfigs, p_mapDataSheetConfigs)
		
		bSuccess = True

		mapPreloadDataMaps = {}

		nDataRowIndex = 0

		nFirstColumnIndex = p_mapDataSheetConfigs.keys()[0]
		nLastColumnIndex  = p_mapDataSheetConfigs.keys()[-1]

		for tplDataSheetRowDataCells in p_cDataSheetDatas :
			mapLineDatas = {}
			
			nDataColumnIndex = nFirstColumnIndex

			for cDataCell in tplDataSheetRowDataCells :
				nDataColumnIndex += 1

				if not xExportHelper.IsDataSheetColumnLanguageAvailable(p_mapDataSheetConfigs[nDataColumnIndex - 1][xConstantData.DATA_SHEET_ROW_LANGUAGE_CODE], p_strExportType, p_mapExportConfigs) :
					continue

				strFieldName = xExportHelper.GetFieldNameAsI18N(p_mapDataSheetConfigs[nDataColumnIndex - 1][xConstantData.DATA_SHEET_ROW_FIELD], p_mapDataSheetConfigs[nDataColumnIndex - 1][xConstantData.DATA_SHEET_ROW_LANGUAGE_CODE], p_strExportType, p_mapExportConfigs)

				mapLineDatas[strFieldName] = cDataCell.value

			lstCategoryLevelValues = []
			
			for nCategoryLevelIndex in lstCategoryLevelIndexs :
				lstCategoryLevelValues.append(tplDataSheetRowDataCells[nCategoryLevelIndex].value)

			lstLineDatas = xWorksheetConvertor.__GetCategoryLevelDataContainer(mapPreloadDataMaps, lstCategoryLevelValues)
			lstLineDatas.append(mapLineDatas)

		if not bSuccess :
			mapPreloadDataMaps.clear()

		return bSuccess, nCategoryLevel, mapPreloadDataMaps
