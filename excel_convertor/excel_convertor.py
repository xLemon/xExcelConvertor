#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import absolute_import
from __future__ import unicode_literals

import codecs
import os
import sys
import traceback
import ConfigParser
import platform

from definitions.constant_data import xConstantData

from utilities.file_utility import xFileUtility
from utilities.excel_helper import xExcelHelper

from core.processor_manager import xProcessorManager
from core.worksheet_convertor import xWorksheetConvertor

from openpyxl.reader.excel import load_workbook

reload(sys)

sys.setdefaultencoding('UTF-8')

__version__      = '0.0.1'
__license__      = 'MIT/Expat'
__author__       = 'lianjin'
__author_qq__    = '55884075'
__author_email__ = 'lianjin.cn@outlook.com'
__url__          = 'https://github.com/xLemon/xExcelConvertor'

def __UniqifyList(p_lstSource) :
	lstSeens  = []
	lstResult = []

	for mixItem in p_lstSource :
		if mixItem in lstSeens :
			continue

		lstSeens.append(mixItem)
		lstResult.append(mixItem)

	return lstResult

def __HasExportType(p_strExportType, p_lstExportTypes) :
	return p_strExportType.upper() in p_lstExportTypes

def __IsValidFile(p_strFileName) :
	strFileSuffix = xFileUtility.GetFileSuffixByName(p_strFileName)

	# 忽略不支持的文件（所支持的文件格式依赖openpyxl库）

	if strFileSuffix not in xConstantData.SUPPORTED_EXCEL_FORMATS :
		return False

	# 忽略临时文件

	if len(p_strFileName) >= 2 and 0 == cmp(p_strFileName[0:2], '~$') :
		return False

	return True

def __FixExportTypes(p_lstExportTypes, p_mapSupporttedTypes) :
	if len(p_lstExportTypes) <= 0 :
		return {}

	lstExportTypes = __UniqifyList(p_lstExportTypes)

	mapExportTypes = {}

	for strExportType in lstExportTypes :
		if p_mapSupporttedTypes.has_key(strExportType) :
			mapExportTypes[strExportType] = p_mapSupporttedTypes[strExportType]

	return mapExportTypes

def __FixPathSeparator(p_strPath) :
	if not isinstance(p_strPath, basestring) or len(p_strPath) <= 0 :
		return ''

	if platform.system().lower() == 'windows' :
		return p_strPath.replace('/', os.sep)
	else :
		return p_strPath.replace('\\', os.sep)

def __FixOptionValue(p_mixValue) :
	if not isinstance(p_mixValue, basestring) or not p_mixValue.strip() :
		return p_mixValue

	lstPaths = []

	if -1 != p_mixValue.find(xConstantData.PATH_SEPARATOR) :
		lstParts = p_mixValue.split(xConstantData.PATH_SEPARATOR)

		for strPart in lstParts :
			if -1 == strPart.find(xConstantData.PATH_IDENTIFIER) and not os.path.isabs(strPart) :
				continue
			else :
				lstPaths.append(strPart)
	elif -1 != p_mixValue.find(xConstantData.PATH_IDENTIFIER) or os.path.isabs(p_mixValue) :
		lstPaths.append(p_mixValue)
	else :
		return p_mixValue

	if len(lstPaths) <= 0 :
		return ''

	for i, strPath in enumerate(lstPaths) :
		if -1 == strPath.find('\\') and -1 == strPath.find('/') :
			continue
		
		strPath = __FixPathSeparator(strPath)

		if -1 != strPath.find(xConstantData.PATH_IDENTIFIER) :
			strPath = strPath.replace(xConstantData.PATH_IDENTIFIER, '{0}{1}'.format(os.path.abspath('.'), os.sep))

		if 0 != cmp(strPath[-1], os.sep) :
			strPath += os.sep
		
		lstPaths[i] = strPath

	return xConstantData.PATH_SEPARATOR.join(lstPaths)

def __SetupOptionValue(p_mapConfigs, p_strKey, p_mixValue) :
	strKey = p_strKey.upper()

	if p_mapConfigs.get(strKey) is not None :
		return

	if isinstance(p_mixValue, bool) :
		p_mapConfigs[strKey] = p_mixValue
	else :
		if p_mixValue.lower() == xConstantData.EXPORT_IDENTIFIER_TRUE or p_mixValue.lower() == xConstantData.EXPORT_IDENTIFIER_YES :
			p_mapConfigs[strKey] = True
		elif p_mixValue.lower() == xConstantData.EXPORT_IDENTIFIER_FALSE or p_mixValue.lower() == xConstantData.EXPORT_IDENTIFIER_NO :
			p_mapConfigs[strKey] = False
		else :
			p_mapConfigs[strKey] = __FixOptionValue(p_mixValue)

def __SetupConfigOptions(p_mapOptions, p_mapConfigs, p_mapGlobalConfigs) :
	for (strKey, mixValue) in p_mapOptions :
		__SetupOptionValue(p_mapConfigs, strKey, mixValue)

	for strKey in p_mapGlobalConfigs :
		__SetupOptionValue(p_mapConfigs, strKey, p_mapGlobalConfigs[strKey])

def __AnalysisConfigs(p_strConfigFilePath) :
	cConfigParser = ConfigParser.ConfigParser()

	cFile = codecs.open(p_strConfigFilePath, 'r', 'UTF-8')

	cConfigParser.readfp(cFile)

	lstSections = cConfigParser.sections()

	if len(lstSections) <= 0 :
		raise Exception('配置文件"{0}"没有指定配置项！'.format(p_strConfigFilePath))

	if not 'GLOBAL' in lstSections :
		raise Exception('配置文件缺少GLOBAL配置项')

	mapExportTypes   = {}
	mapGlobalConfigs = {}
	mapExportConfigs = {}

	__SetupConfigOptions(cConfigParser.items('GLOBAL'), mapGlobalConfigs, {})

	mapExportConfigs['EXCEL_DIRECTORY'] = mapGlobalConfigs['EXCEL_DIRECTORY']
	mapExportConfigs['DEEP_PROCESS']    = mapGlobalConfigs['DEEP_PROCESS']

	for strSection in lstSections :
		if strSection == 'GLOBAL' :
			continue

		if strSection == 'COPYRIGHT' :
			if mapExportConfigs.get(strSection.upper()) is None :
				mapExportConfigs[strSection.upper()] = {}

			# 版权信息不需要合并全局配置

			__SetupConfigOptions(cConfigParser.items(strSection), mapExportConfigs[strSection.upper()], {})
			
			continue

		if len(strSection) < 7 or 0 != cmp(strSection[0:7], 'EXPORT_') :
			continue

		strExportType = strSection[7:].upper()

		if mapExportConfigs.get('EXPORTS') is None :
			mapExportConfigs['EXPORTS'] = {}

		if mapExportConfigs['EXPORTS'].get(strExportType) is None :
			mapExportConfigs['EXPORTS'][strExportType] = {}

		# 解析导出类型配置，若导出类型配置与全局配置有共同的配置项，以导出类型配置优先

		__SetupConfigOptions(cConfigParser.items(strSection), mapExportConfigs['EXPORTS'][strExportType], mapGlobalConfigs)

		# 删除不必要的key

		if mapExportConfigs['EXPORTS'][strExportType].get('EXCEL_DIRECTORY') is not None :
			mapExportConfigs['EXPORTS'][strExportType].pop('EXCEL_DIRECTORY')

		bIgnore = False

		if mapExportConfigs['EXPORTS'][strExportType].get('EXPORT_PROCESSOR') is None or len(mapExportConfigs['EXPORTS'][strExportType]['EXPORT_PROCESSOR']) <= 0 :
			print('{0} : 必须配置导出处理器'.format(strExportType))
			bIgnore = True
		else :
			if xProcessorManager.IsProcessorExist(mapExportConfigs['EXPORTS'][strExportType]['EXPORT_PROCESSOR']) :
				mapExportTypes[strExportType] = mapExportConfigs['EXPORTS'][strExportType]['EXPORT_PROCESSOR'].lower()
			else :
				print('跳过不支持的导出格式 : {0}，若要支持该格式，请在{1}目录中扩展'.format(strExportType, xProcessorManager.GetProcessorPath()))
				bIgnore = True

		if bIgnore :
			if mapExportConfigs['EXPORTS'].get(strExportType) is not None :
				mapExportConfigs['EXPORTS'].pop(strExportType)
			continue

		if mapExportConfigs['EXPORTS'][strExportType].get('EXPORT_SUFFIX') is None or len(mapExportConfigs['EXPORTS'][strExportType]['EXPORT_SUFFIX']) <= 0 :
			mapExportConfigs['EXPORTS'][strExportType]['EXPORT_SUFFIX'] = mapExportConfigs['EXPORTS'][strExportType]['EXPORT_PROCESSOR'].lower()

		if mapExportConfigs['EXPORTS'][strExportType]['EXPORT_PROCESSOR'].upper() != 'SQL' :
			continue

		if mapExportConfigs['EXPORTS'][strExportType].get('FORMAT_DATA') is not None :
			mapExportConfigs['EXPORTS'][strExportType].pop('FORMAT_DATA')

		if mapExportConfigs['EXPORTS'][strExportType].get('DATA_NAMESPACE') is not None :
			mapExportConfigs['EXPORTS'][strExportType].pop('DATA_NAMESPACE')

	cFile.close()

	return mapExportConfigs, mapExportTypes

def __AnalysisIndexSheetConfigs(p_cIndexSheet, p_cCell, p_nRowIndex, p_mapExportTypes) :
	if p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_ENABLE].value is None :
		return True, True, {}, None # 没有配置导出项，跳过

	if p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_NAME].value is None :
		return False, False, {}, '错误：工作表 [{0}] [{1}{2}] 单元格未配置!'.format(p_cIndexSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(xConstantData.INDEX_SHEET_DATA_COLUMN_NAME), p_nRowIndex)

	if p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_ALIAS].value is None :
		return False, False, {}, '错误：工作表 [{0}] [{1}{2}] 单元格未配置!'.format(p_cIndexSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(xConstantData.INDEX_SHEET_DATA_COLUMN_ALIAS), p_nRowIndex)

	bExportEmptyDataItem = True

	strExportEmptyDataItemValue = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_EXPORT_EMPTY_DATA_ITEM].value

	if strExportEmptyDataItemValue is not None :
		if strExportEmptyDataItemValue.lower() == xConstantData.EXPORT_IDENTIFIER_TRUE or strExportEmptyDataItemValue.lower() == xConstantData.EXPORT_IDENTIFIER_YES :
			bExportEmptyDataItem = True
		elif strExportEmptyDataItemValue.lower() == xConstantData.EXPORT_IDENTIFIER_FALSE or strExportEmptyDataItemValue.lower() == xConstantData.EXPORT_IDENTIFIER_NO :
			bExportEmptyDataItem = False
		else :
			bExportEmptyDataItem = True

	mapIndexSheetConfigs = {}

	mapIndexSheetConfigs['DATA_SHEET']              = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_NAME].value
	mapIndexSheetConfigs['DATA_FILE_NAME']          = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_ALIAS].value
	mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES'] = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_ENABLE].value.upper().split(',')
	mapIndexSheetConfigs['EXPORT_EMPTY_DATA_ITEM']  = bExportEmptyDataItem

	if len(mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES']) <= 0 :
		return False, {}, '错误：工作表 [{0}] [{1}{2}] 单元格导出类型配置错误!'.format(p_cIndexSheet.title, xExcelHelper.ConvertColumnIndexToColumnName(xConstantData.INDEX_SHEET_DATA_COLUMN_ENABLE), p_nRowIndex)

	mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES'] = __FixExportTypes(mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES'], p_mapExportTypes)

	if p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_TABLE_ENGINE].value is None :
		mapIndexSheetConfigs['DATA_TABLE_ENGINE'] = xConstantData.DEFAULT_DATA_TABLE_ENGINE
	else:
		mapIndexSheetConfigs['DATA_TABLE_ENGINE'] = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_TABLE_ENGINE].value

	if p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_TABLE_CHARSET].value is None :
		mapIndexSheetConfigs['DATA_TABLE_CHARSET'] = xConstantData.DEFAULT_DATA_TABLE_CHARSET
	else:
		mapIndexSheetConfigs['DATA_TABLE_CHARSET'] = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_TABLE_CHARSET].value

	if p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_TABLE_COLLATE].value is None :
		mapIndexSheetConfigs['DATA_TABLE_COLLATE'] = xConstantData.DEFAULT_DATA_TABLE_COLLATE
	else:
		mapIndexSheetConfigs['DATA_TABLE_COLLATE'] = p_cCell[xConstantData.INDEX_SHEET_DATA_COLUMN_TABLE_COLLATE].value

	return False, True, mapIndexSheetConfigs, None

def __PrepareExportDirectories(p_mapExportConfigs, p_lstExportTypes) :
	for strExportType in p_lstExportTypes :
		mapExportConfig = p_mapExportConfigs['EXPORTS'].get(strExportType)

		if mapExportConfig is None :
			continue

		strExportDirectory = mapExportConfig['EXPORT_DIRECTORY']

		if xFileUtility.IsFileOrDirectoryExist(strExportDirectory) :
			continue

		xFileUtility.CreateDirectory(strExportDirectory)

def __ProcessGlobalFile(p_mapExportTypes, p_mapProcessMarks, p_strWorkbookName, p_mapExportConfigs, p_mapDatabaseConfigs) :
	if len(p_mapExportTypes) <= 0 :
		return

	for (strExportType, strExportProcessor) in p_mapExportTypes.items() :
		if p_mapProcessMarks.get(strExportType) is None :
			p_mapProcessMarks[strExportType] = False

		if p_mapProcessMarks[strExportType] :
			continue

		cProcessorInstance = xProcessorManager.GetProcessorInstance(strExportProcessor, p_mapExportConfigs['EXPORTS'][strExportType]['EXPORT_SUFFIX'], strExportType)

		p_mapProcessMarks[strExportType] = cProcessorInstance.ProcessGlobalFile(p_strWorkbookName, p_mapExportConfigs, p_mapDatabaseConfigs)

def __ProcessDump(p_mapExportConfigs, p_mapExportTypes) :
	setDirectoryFiles = xFileUtility.GetDirectoryFiles(p_mapExportConfigs['EXCEL_DIRECTORY'], p_mapExportConfigs['DEEP_PROCESS'])

	for strFilePath in setDirectoryFiles :
		strFileName = xFileUtility.GetFileNameFromPath(strFilePath)

		if not __IsValidFile(strFileName) :
			continue

		print('正在加载工作簿 : {0}'.format(strFileName))

		cWorkbook = load_workbook(filename = strFilePath, read_only = False, data_only = True)

		strWorkbookName = xFileUtility.GetFileNameFromPath(strFilePath, True)

		lstSheetNames = cWorkbook.get_sheet_names()

		if len(lstSheetNames) <= 0 :
			continue # 工作簿没有配置工作表，跳过

		# 列表目录的Sheet必须在第一个（Sheet的命名随意）

		cIndexSheet = cWorkbook.get_sheet_by_name(lstSheetNames[0])

		nIgnoreRows = 2
		nRowIndex   = 1

		if not cIndexSheet.max_row > nIgnoreRows :
			continue # 工作簿索引页无配置项，跳过

		print('正在处理工作簿 : {0}'.format(strFilePath))

		mapProcessMarks    = {}
		mapDatabaseConfigs = {}

		cRowCells = cIndexSheet.get_squared_range(1, 1, cIndexSheet.max_column, cIndexSheet.max_row)

		for cCell in cRowCells :
			if 1 == nRowIndex :
				mapDatabaseConfigs['DATABASE_NAME']     = cCell[xConstantData.INDEX_SHEET_GLOBAL_COLUMN_DATABASE_NAME].value     # 数据库名
				mapDatabaseConfigs['DATABASE_CHARSET']  = cCell[xConstantData.INDEX_SHEET_GLOBAL_COLUMN_DATABASE_CHARSET].value  # 数据库字符集
				mapDatabaseConfigs['DATABASE_COLLATE']  = cCell[xConstantData.INDEX_SHEET_GLOBAL_COLUMN_DATABASE_COLLATE].value  # 数据库校对规则
				mapDatabaseConfigs['DATA_TABLE_PREFIX'] = cCell[xConstantData.INDEX_SHEET_GLOBAL_COLUMN_DATA_TABLE_PREFIX].value # 数据表前缀

				if mapDatabaseConfigs['DATABASE_NAME'] is None :
					mapDatabaseConfigs['DATABASE_NAME'] = ''

				if mapDatabaseConfigs['DATA_TABLE_PREFIX'] is None :
					mapDatabaseConfigs['DATA_TABLE_PREFIX'] = ''

			nRowIndex += 1

			if (nRowIndex - 1) <= nIgnoreRows :
				continue

			bIgnore, bSuccess, mapIndexSheetConfigs, strRaiseMessage = __AnalysisIndexSheetConfigs(cIndexSheet, cCell, nRowIndex - 1, p_mapExportTypes)

			if bIgnore :
				continue

			if not bSuccess :
				raise Exception(strRaiseMessage)

			# 仅创建实际导出的类型所对应的目录

			# __PrepareExportDirectories(p_mapExportConfigs, mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES'])

			# 有些类型的文件（例如SQL），它所有导出的数据需要存放在同一份文件里，在此处处理这类需求

			__ProcessGlobalFile(mapIndexSheetConfigs['DATA_SHEET_EXPORT_TYPES'], mapProcessMarks, strWorkbookName, p_mapExportConfigs, mapDatabaseConfigs)

			if not xWorksheetConvertor.ConvertWorksheet(strWorkbookName, cWorkbook, p_mapExportConfigs, mapDatabaseConfigs, mapIndexSheetConfigs) :
				raise Exception('发生错误，退出！')

def main() :
	strConfigFile = os.path.join(os.path.abspath('.'), 'config', 'export_config.ini')

	if not xFileUtility.IsFileOrDirectoryExist(strConfigFile) :
		raise Exception('配置文件"{0}"不存在！'.format(strConfigFile))

	mapExportConfigs, mapExportTypes = __AnalysisConfigs(strConfigFile)

	__ProcessDump(mapExportConfigs, mapExportTypes)

if __name__ == '__main__' :
	try :
		main()
	except Exception as cException :
		traceback.print_exc()
		sys.exit(1)
